#!/usr/bin/env python3
"""
xlsx-to-sql.py — แปลงฐานข้อมูลเดิม (Google Sheets export: TapTime.xlsx)
                 เป็นไฟล์ SQL สำหรับ import เข้า Supabase / PostgreSQL

  usage:  python3 scripts/xlsx-to-sql.py ../TapTime.xlsx supabase/seed.sql

หมายเหตุการแมป
  ชีต Employees        -> ตาราง profiles
  ชีต LeaveRequests    -> ตาราง leave_requests
  ชีต TimeEditRequests -> ตาราง time_edit_requests
  ชีต LeaveTypes       -> ตาราง leave_types
  คอลัมน์คงชื่อเดิม (camelCase) จึงถูก quote ด้วย " เสมอ
"""
import sys
import datetime
import openpyxl

SHEET_TABLE = {
    "Config":           "config",
    "Branches":         "branches",
    "Positions":        "positions",
    "LeaveTypes":       "leave_types",
    "Employees":        "profiles",
    "Attendance":       "attendance",
    "LeaveRequests":    "leave_requests",
    "Holidays":         "holidays",
    "TimeEditRequests": "time_edit_requests",
}

# ลำดับ insert ให้เคารพ foreign key
ORDER = ["Config", "Branches", "Positions", "LeaveTypes", "Employees",
         "Attendance", "LeaveRequests", "Holidays", "TimeEditRequests"]

# คอลัมน์ที่ต้องเป็น 'HH:mm' (text) — ชีตมักเก็บเป็น time/Date
TIME_COLS = {"workStart", "workEnd", "checkInTime", "checkOutTime",
             "oldCheckIn", "oldCheckOut", "newCheckIn", "newCheckOut"}
# คอลัมน์ที่เป็นวันที่ (date)
DATE_COLS = {"date", "startDate", "endDate", "birthDate", "expireDate"}
# คอลัมน์ที่เป็น timestamp ข้อความ 'yyyy-MM-dd HH:mm'
STAMP_COLS = {"requestedAt", "decidedAt"}
# คอลัมน์ที่ต้องเป็น text แม้ค่าจะดูเป็นตัวเลข (เช่น code '00000', postcode)
FORCE_TEXT = {"code", "postcode", "phone", "nationalId", "value", "token"}
BOOL_COLS = {"isHead"}


def q(s):
    """escape ค่าเป็น string literal ของ SQL"""
    return "'" + str(s).replace("'", "''") + "'"


def conv(col, v):
    if v is None or v == "":
        return "NULL"

    if col in BOOL_COLS:
        return "true" if v in (True, 1, "TRUE", "true", "1") else "false"

    if col in TIME_COLS:
        if isinstance(v, datetime.time):
            return q("%02d:%02d" % (v.hour, v.minute))
        if isinstance(v, datetime.datetime):
            return q(v.strftime("%H:%M"))
        return q(str(v)[:5])

    if col in DATE_COLS:
        if isinstance(v, (datetime.datetime, datetime.date)):
            return q(v.strftime("%Y-%m-%d"))
        return q(str(v)[:10])

    if col in STAMP_COLS:
        if isinstance(v, datetime.datetime):
            return q(v.strftime("%Y-%m-%d %H:%M"))
        return q(str(v))

    if col in FORCE_TEXT:
        # เลี่ยง 0.0 / 12130.0 ที่ Sheets แปลงเป็น float
        if isinstance(v, float) and v.is_integer():
            return q(str(int(v)))
        return q(str(v))

    if isinstance(v, bool):
        return "true" if v else "false"
    if isinstance(v, (int, float)):
        return repr(v)
    if isinstance(v, (datetime.datetime, datetime.date)):
        return q(v.strftime("%Y-%m-%d"))
    return q(v)


def main(src, dst):
    wb = openpyxl.load_workbook(src, data_only=True)
    out = [
        "-- ============================================================",
        "-- TapTime — seed data (แปลงอัตโนมัติจาก %s)" % src.split("/")[-1],
        "-- สร้างโดย scripts/xlsx-to-sql.py — รันหลัง schema.sql",
        "-- ============================================================",
        "",
    ]

    for sheet_name in ORDER:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        table = SHEET_TABLE[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header = [h for h in rows[0] if h]
        body = [r for r in rows[1:] if r and r[0] not in (None, "")]

        out.append("-- ---- %s (%d แถว) --------------------------------"
                   % (table, len(body)))
        if not body:
            out.append("-- (ไม่มีข้อมูล)\n")
            continue

        cols = ", ".join('"%s"' % c for c in header)
        values = []
        for r in body:
            vals = [conv(header[i], r[i] if i < len(r) else None)
                    for i in range(len(header))]
            values.append("  (" + ", ".join(vals) + ")")

        pk = header[0]
        out.append("insert into %s (%s) values" % (table, cols))
        out.append(",\n".join(values))
        out.append('on conflict ("%s") do nothing;\n' % pk)

    # weeklyOff ในชีตอาจถูกเก็บเป็นตัวเลข (0) ทำให้ค่ากลายเป็น '0.0'
    out.append("-- ---- ทำความสะอาดค่า config ที่ Sheets แปลงเป็น float ----")
    out.append("update config set \"value\" = rtrim(rtrim(\"value\", '0'), '.')")
    out.append("  where \"value\" ~ '^[0-9]+\\.0+$';")
    out.append("")

    with open(dst, "w", encoding="utf-8") as f:
        f.write("\n".join(out))
    print("เขียน %s เรียบร้อย" % dst)


if __name__ == "__main__":
    src = sys.argv[1] if len(sys.argv) > 1 else "../TapTime.xlsx"
    dst = sys.argv[2] if len(sys.argv) > 2 else "supabase/seed.sql"
    main(src, dst)
